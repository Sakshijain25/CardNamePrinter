

import openpyxl
from openpyxl.styles import PatternFill
import json

def highlight_matching_entries(file_path, sheet_name, column_index, match_list, highlight_color):
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Define the fill style for highlighting
    fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type="solid")

    # Iterate over the specified column and highlight matching entries
    for row in sheet.iter_rows():
            if row[column_index].value in match_list:
                for cell in row:
                    cell.fill = fill
    # Save the modified workbook
    workbook.save(file_path)

# Example usage
file_path = "D:\Sakshi\weddinginvitation\input\Cards List.xlsx"  # Replace with the desired output path

sheet_name = "Sheet2"
column_index = 7  # Assuming the column index is 1-based (A=1, B=2, etc.)
contacts_path = r"D:\Sakshi\weddinginvitation\input\not_found.json"  # Replace with the desired output path


with open(contacts_path, 'r') as json_file:
    match_list = json.load(json_file)

highlight_color = "FFFF00"  # Yellow

highlight_matching_entries(file_path, sheet_name, column_index, match_list, highlight_color)
